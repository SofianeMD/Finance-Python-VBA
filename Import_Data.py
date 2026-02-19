import yfinance as yf
import pandas as pd
from openpyxl import load_workbook
import os

FILE_NAME = "/Users/sofiane/Desktop/Analyse_Portefeuille.xlsm"
SHEET_NAME = "Data_Finance"
TK = ["AAPL", "NVDA"]

def update_excel():
    try:
       
        df = yf.download(TK, period="60d")['Close'].dropna()
        
        if not os.path.exists(FILE_NAME):
            print("Le fichier Excel n'a pas été trouvé.")
            return

        book = load_workbook(FILE_NAME, keep_vba=True)
        
        with pd.ExcelWriter(FILE_NAME, engine='openpyxl') as writer:
            writer.book = book

            if SHEET_NAME in book.sheetnames:
                book.remove(book[SHEET_NAME])
            df.to_excel(writer, sheet_name=SHEET_NAME)
            
        print("Mise à jour terminée.)")
        
    except Exception as e:
        print(f"Erreur : {e}")

if __name__ == "__main__":
    update_excel()
